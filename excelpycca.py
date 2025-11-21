import fitz  
import pandas as pd
import os
from openpyxl import load_workbook

# Rutas
pdf_path = r"C:\Users\Umco\OneDrive\Desktop\ETIQUETAS DEPRATI\PYCCA\168706.pdf"
output_excel = r"C:\Users\Umco\OneDrive\Desktop\ETIQUETAS DEPRATI\PYCCA\etiquetas_pycca.xlsx"

# Extraer texto del PDF
doc = fitz.open(pdf_path)
lineas = []

for page in doc:
    bloques = page.get_text("blocks")
    for b in bloques:
        texto = b[4].strip()
        if texto and texto.lower() != "incluido iva":
            lineas.append(texto)

# Agrupar cada 3 líneas
productos = []
for i in range(0, len(lineas), 3):
    try:
        codigo = lineas[i]
        precio_raw = lineas[i + 1].replace("$", "").replace(".", "").strip()

        if len(precio_raw) >= 3:
            precio_entero = precio_raw[:-2]
            precio_decimal = precio_raw[-2:]
        elif len(precio_raw) == 2:
            precio_entero = "0"
            precio_decimal = precio_raw
        elif len(precio_raw) == 1:
            precio_entero = "0"
            precio_decimal = f"0{precio_raw}"
        else:
            precio_entero = "0"
            precio_decimal = "00"

        descripcion = lineas[i + 2]
        productos.append({
            "Código": codigo,
            "Precio_Entero": precio_entero,
            "Precio_Decimal": precio_decimal,
            "Descripción": descripcion,
            #"Cantidad": 1,
        })
    except IndexError:
        print(f" Entrada incompleta en bloque {i}, se omite.")

# Convertir a DataFrame
df = pd.DataFrame(productos)

# Cargar libro y hoja
wb = load_workbook(output_excel)
ws = wb["Precios"]

# Borrar columnas A a E desde la fila 2
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        if cell.column_letter in ["A", "B", "C", "D"]: # "E" en el caso de que no este funcionando
            cell.value = None

# Insertar nuevos datos desde fila 2
for i, fila in enumerate(df.values.tolist(), start=2):
    for j, valor in enumerate(fila, start=1):  # Columnas A=1 a E=5
        ws.cell(row=i, column=j, value=valor)

# Guardar cambios
wb.save(output_excel)

print(f"\nExcel actualizado {output_excel}")
